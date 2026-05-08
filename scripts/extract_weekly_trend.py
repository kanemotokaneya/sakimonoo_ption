#!/usr/bin/env python3
"""Extract multi-week futures OI trend by participant.

Reads all `*_indexfut_oi_by_tp.xlsx` files in `data/` (sorted by YYYYMMDD prefix),
extracts per-participant net OI per market (大/mini/TOPIX) for the nearest major
limgetsu, computes WoW deltas across weeks, writes `data/weekly_trend.json`.

Designed to run AFTER extract.py in the pipeline. Independent module — does not
modify extract.py state. Safe to run repeatedly.
"""
import os
import re
import json
import glob
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print('[weekly_trend] ERROR: openpyxl not installed', file=sys.stderr)
    sys.exit(1)


# --- Customer type classification (mirrors extract.py CUSTOMER_TYPES) ---
# Tuples ordered by priority: more specific keywords FIRST so that a name
# containing "モルガンＭＵＦＧ" matches CTA before plain "モルガン".
CUSTOMER_TYPES = [
    ('ゴールドマン',      'グローバルマクロ'),
    ('ＪＰモルガン',       'グローバルマクロ'),
    ('シティグループ',     'グローバルマクロ'),
    ('ビーオブエー',       '長期投資志向'),
    ('モルガンＭＵＦＧ',  'トレンドフォロー（CTA）'),
    ('ＡＢＮ',             'アービトラージ（裁定取引）'),
    ('ソシエテ',           'アービトラージ（裁定取引）'),
    ('ＢＮＰ',             'アービトラージ（裁定取引）'),
    ('野村',               '国内機関投資家'),
    ('みずほ',             '国内機関投資家'),
    ('ＳＢＩ',             '国内個人投資家（ネットトレーダー）'),
    ('楽天',               '国内個人投資家（ネットトレーダー）'),
    ('松井',               '国内個人投資家（ネットトレーダー）'),
    ('マネックス',         '国内個人投資家（ネットトレーダー）'),
    ('三菱ＵＦＪｅスマート','国内個人投資家（ネットトレーダー）'),
]


def classify(name):
    """Return customer-type label or '-' (the "-" bucket from screenshot 6454)."""
    if not name:
        return '-'
    for kw, label in CUSTOMER_TYPES:
        if kw in name:
            return label
    return '-'


# --- Section parsing -------------------------------------------------------

SECTION_KEYS = [
    ('n225_large', '日経225先物', 'mini'),   # exclude rows containing 'mini'
    ('n225_mini',  '日経225mini', None),
    ('topix',      'TOPIX先物',    None),
]


def find_sections(ws):
    """Return {section_key: (start_row, end_row_exclusive)}.

    end_row_exclusive is the row of the NEXT section header, or max_row+1.
    """
    bounds = OrderedDict()
    starts = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a:
            continue
        s = str(a)
        if '＜' not in s:
            continue
        for key, kw, exclude in SECTION_KEYS:
            if kw in s and (not exclude or exclude not in s):
                starts.append((key, r))
                break

    for i, (key, start) in enumerate(starts):
        end = starts[i + 1][1] if i + 1 < len(starts) else ws.max_row + 1
        bounds[key] = (start, end)
    return bounds


def extract_section_nets(ws, start_row, end_row):
    """For the FIRST limgetsu block of a section, build {broker: net_oi}.

    Columns (1-indexed) within the first limgetsu block:
      D(4) = sell name  / E(5) = sell qty
      G(7) = buy  name  / H(8) = buy  qty

    The 2nd limgetsu block (cols K-R) is intentionally ignored — we only
    track the nearest major limgetsu (06月限 in this snapshot).
    """
    nets = {}
    for r in range(start_row + 2, end_row):  # +2 skips section header + col headers
        sell_name = ws.cell(r, 4).value
        sell_qty  = ws.cell(r, 5).value
        buy_name  = ws.cell(r, 7).value
        buy_qty   = ws.cell(r, 8).value

        if sell_name and isinstance(sell_qty, (int, float)) and sell_qty > 0:
            n = str(sell_name).strip()
            nets[n] = nets.get(n, 0.0) - float(sell_qty)
        if buy_name and isinstance(buy_qty, (int, float)) and buy_qty > 0:
            n = str(buy_name).strip()
            nets[n] = nets.get(n, 0.0) + float(buy_qty)
    return nets


def detect_first_limgetsu(ws, section_start_row):
    """Return string label of the 1st limgetsu (e.g. '2026年06月限月'),
    or empty string if not found."""
    # The limgetsu label sits in column B (index 2) of the first data row
    # under the section header.
    for r in range(section_start_row, section_start_row + 4):
        v = ws.cell(r, 2).value
        if v and '限月' in str(v):
            return str(v).strip()
    return ''


# --- Main weekly aggregation ----------------------------------------------

def extract_weekly_trend(data_dir, max_weeks=5):
    """Build the multi-week trend dict.

    Returns:
      {
        'weeks': ['20260410', '20260417', ...],
        'limgetsu': {'n225_large': '2026年06月限月', ...},
        'sections': {
          'n225_large': [
             {'broker': 'ゴールドマン証券', 'category': 'グローバルマクロ',
              'oi_history': [..., -117], 'wow': [3378, ..., -117], 'oi_current': -117},
             ...
          ],
          'n225_mini':  [...],
          'topix':      [...],
        },
        'generated_at': '2026-05-08',
      }

    On no files found, returns {'error': '...'}.
    """
    pattern = os.path.join(data_dir, '*_indexfut_oi_by_tp.xlsx')
    candidates = []
    for fp in glob.glob(pattern):
        m = re.search(r'(\d{8})', os.path.basename(fp))
        if m:
            candidates.append((m.group(1), fp))

    if not candidates:
        return {'error': 'No *_indexfut_oi_by_tp.xlsx files found in %s' % data_dir}

    candidates.sort(key=lambda x: x[0])  # ascending date
    candidates = candidates[-max_weeks:]
    weeks = [d for d, _ in candidates]
    print('[weekly_trend] using %d weekly files: %s' % (len(weeks), weeks))

    # week_data[date_str][section_key][broker] = net_oi
    week_data = {}
    limgetsu_per_section = {}

    for date_str, fp in candidates:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except Exception as e:
            print('[weekly_trend] WARN: failed to open %s: %s' % (fp, e))
            continue
        ws = wb[wb.sheetnames[0]]
        sections = find_sections(ws)
        week_data[date_str] = {}
        for sec_key, (s, e) in sections.items():
            week_data[date_str][sec_key] = extract_section_nets(ws, s, e)
            if sec_key not in limgetsu_per_section:
                lim = detect_first_limgetsu(ws, s)
                if lim:
                    limgetsu_per_section[sec_key] = lim

    # Build per-section participant rows
    output = {
        'weeks': weeks,
        'limgetsu': limgetsu_per_section,
        'sections': {},
        'generated_at': '',
    }

    for sec_key in ('n225_large', 'n225_mini', 'topix'):
        # Collect all brokers seen in any week for this section
        all_brokers = set()
        for d in weeks:
            all_brokers.update(week_data.get(d, {}).get(sec_key, {}).keys())

        rows = []
        for broker in all_brokers:
            oi_hist = []
            for d in weeks:
                oi_hist.append(week_data.get(d, {}).get(sec_key, {}).get(broker, 0.0))
            # WoW deltas: wow[i] = oi_hist[i] - oi_hist[i-1] (i>=1)
            # For i=0 (first week), no prior → use 0 (or could mark None)
            wow = []
            for i in range(len(oi_hist)):
                if i == 0:
                    wow.append(None)  # no prior week → None means "no data"
                else:
                    wow.append(oi_hist[i] - oi_hist[i - 1])
            rows.append({
                'broker': broker,
                'category': classify(broker),
                'oi_history': oi_hist,
                'wow': wow,
                'oi_current': oi_hist[-1] if oi_hist else 0.0,
            })

        # Sort: by category order, then by abs(oi_current) descending
        cat_order = {label: i for i, (_, label) in enumerate(CUSTOMER_TYPES)}
        cat_order_default = {
            'グローバルマクロ': 0,
            '長期投資志向': 1,
            'トレンドフォロー（CTA）': 2,
            'アービトラージ（裁定取引）': 3,
            '国内機関投資家': 4,
            '国内個人投資家（ネットトレーダー）': 5,
            '-': 99,
        }
        rows.sort(key=lambda r: (
            cat_order_default.get(r['category'], 99),
            -abs(r['oi_current']),
        ))
        output['sections'][sec_key] = rows

    # Stamp generated date from latest week
    output['generated_at'] = weeks[-1] if weeks else ''
    return output


# --- CLI entry point -------------------------------------------------------

def main():
    import argparse
    ap = argparse.ArgumentParser(description='Extract weekly futures OI trend')
    ap.add_argument('--data-dir', default='data', help='dir with *_indexfut_oi_by_tp.xlsx')
    ap.add_argument('--out',      default='data/weekly_trend.json', help='output JSON path')
    ap.add_argument('--max-weeks', type=int, default=5)
    args = ap.parse_args()

    result = extract_weekly_trend(args.data_dir, max_weeks=args.max_weeks)

    if 'error' in result:
        print('[weekly_trend] %s' % result['error'])
        # still write a stub so the page can show "データ蓄積中"
        os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
        with open(args.out, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        return 0

    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    n_total = sum(len(rows) for rows in result['sections'].values())
    print('[weekly_trend] wrote %s | %d weeks | %d total rows across 3 sections'
          % (args.out, len(result['weeks']), n_total))
    return 0


if __name__ == '__main__':
    sys.exit(main())
