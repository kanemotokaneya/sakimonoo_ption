#!/usr/bin/env python3
"""Aggregate participant volume across ALL four session files, by contract.

The daily J-NET file only shows off-auction blocks. Three more files exist —
day auction, night auction, night J-NET — and together they show how much of a
strike's activity came from large off-auction blocks versus ordinary on-screen
flow, and which brokers were present in each.

Output (per expiry/side/strike):
    total     total lots across all sessions
    by_venue  {'day_auction':n,'day_jnet':n,'night_auction':n,'night_jnet':n}
    brokers   [{'broker','vol','cat','venues'}]  top participants
    block_pct share of volume that came through J-NET (off-auction blocks)

This is the "who was in this strike, and through which door" layer that sits
between the raw OI change and the IV reaction.
"""
import argparse
import collections
import glob
import json
import os
import re

import openpyxl

# venue key -> (filename marker, label)
VENUES = [
    ('day_auction', '_volume_by_participant_whole_day.xlsx', '日中・立会'),
    ('day_jnet', '_volume_by_participant_whole_day_J-NET.xlsx', '日中・J-NET'),
    ('night_auction', '_volume_by_participant_night.xlsx', '夜間・立会'),
    ('night_jnet', '_volume_by_participant_night_J-NET.xlsx', '夜間・J-NET'),
]

_DOM = ('野村', 'みずほ', '大和', '三菱', 'ＳＭＢＣ', '日興', '東海', 'ＳＢＩ',
        '楽天', '松井', 'マネックス', '岡三', '光世', '日産', 'auカブコム')
_HF = ('ＡＢＮ', 'サスケハナ', 'インタラクティブ')
_US = ('ゴールドマン', 'ＪＰモルガン', 'モルガン', 'シティ', 'ビーオブエー',
       'メリル', 'ジェフリーズ')
_EU = ('ＵＢＳ', 'ＢＮＰ', 'ソシエテ', 'バークレイズ', 'ドイツ', 'ＨＳＢＣ',
       'クレディ', 'ナティクシス')


def _classify(b):
    s = str(b or '')
    for k in _DOM:
        if k in s:
            return 'domestic'
    for k in _HF:
        if k in s:
            return 'hf'
    for k in _US:
        if k in s:
            return 'us'
    for k in _EU:
        if k in s:
            return 'eu'
    return 'other'


def _parse_contract(prod):
    """'NIKKEI 225 OOP C2609-67000' -> ('C','2609',67000); futures -> None."""
    m = re.search(r'([CP])(\d{4})-(\d+)', str(prod or ''))
    if not m:
        return None
    return m.group(1), m.group(2), int(m.group(3))


def _read_one(path):
    """-> list of (side, expiry, strike, broker, vol) for option rows."""
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['手口上位一覧'] if '手口上位一覧' in wb.sheetnames else wb.worksheets[0]
    for r in ws.iter_rows(values_only=True):
        if not r or str(r[0]).strip() != 'NK225E':
            continue
        c = _parse_contract(r[2] if len(r) > 2 else None)
        if not c:
            continue
        broker = str(r[5]).strip() if len(r) > 5 and r[5] else None
        vol = r[7] if len(r) > 7 and isinstance(r[7], (int, float)) else None
        if broker and vol:
            out.append((c[0], c[1], c[2], broker, float(vol)))
    return out


def build_venue_flow(data_dir, date=None):
    """Aggregate all available session files for the latest (or given) date."""
    found = {}
    for key, marker, label in VENUES:
        pat = os.path.join(data_dir, '*' + marker)
        files = sorted(glob.glob(pat))
        # day_auction's marker is a suffix of day_jnet's name? No — J-NET has
        # the extra suffix, so filter it out explicitly.
        if key == 'day_auction':
            files = [f for f in files if 'J-NET' not in os.path.basename(f)]
        if key == 'night_auction':
            files = [f for f in files if 'J-NET' not in os.path.basename(f)]
        if date:
            files = [f for f in files if date in os.path.basename(f)]
        if files:
            found[key] = files[-1]

    if not found:
        return None

    # infer date from any filename
    d = date
    if not d:
        m = re.search(r'(\d{8})', os.path.basename(list(found.values())[0]))
        d = m.group(1) if m else ''

    # (expiry, side, strike) -> aggregation
    rows = collections.defaultdict(lambda: {
        'total': 0.0,
        'by_venue': collections.defaultdict(float),
        'brokers': collections.defaultdict(lambda: {'vol': 0.0, 'venues': set()}),
    })
    venue_labels = {k: lbl for k, _, lbl in VENUES}
    for key, path in found.items():
        try:
            for side, exp, strike, broker, vol in _read_one(path):
                e = rows[(exp, side, strike)]
                e['total'] += vol
                e['by_venue'][key] += vol
                b = e['brokers'][broker]
                b['vol'] += vol
                b['venues'].add(venue_labels[key])
        except Exception as ex:
            print('[venue_flow] %s parse failed: %s' % (key, ex))

    out = []
    for (exp, side, strike), e in rows.items():
        jnet = e['by_venue'].get('day_jnet', 0) + e['by_venue'].get('night_jnet', 0)
        brokers = sorted(
            ({'broker': b, 'vol': round(v['vol']), 'cat': _classify(b),
              'venues': sorted(v['venues'])}
             for b, v in e['brokers'].items()),
            key=lambda x: -x['vol'])[:6]
        out.append({
            'expiry': exp, 'side': side, 'strike': strike,
            'total': round(e['total']),
            'by_venue': {k: round(v) for k, v in e['by_venue'].items()},
            'block_pct': round(jnet / e['total'] * 100) if e['total'] else 0,
            'brokers': brokers,
        })
    out.sort(key=lambda x: -x['total'])
    return {'date': d, 'venues_found': sorted(found.keys()), 'rows': out}


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--data-dir', default='data')
    ap.add_argument('--out', default='data/venue_flow.json')
    ap.add_argument('--date', default=None)
    a = ap.parse_args()
    res = build_venue_flow(a.data_dir, a.date)
    if res is None:
        print('[venue_flow] no session files found — skipped')
    else:
        json.dump(res, open(a.out, 'w', encoding='utf-8'),
                  ensure_ascii=False, separators=(',', ':'))
        print('[venue_flow] %s: %d contracts across %s -> %s'
              % (res['date'], len(res['rows']), res['venues_found'], a.out))
