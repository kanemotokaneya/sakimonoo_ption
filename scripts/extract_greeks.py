#!/usr/bin/env python3
"""
extract_greeks.py — Per-strike Black-Scholes greeks + OI-weighted Gamma
Exposure (GEX) for the JPX Nikkei 225 options board.

Inputs (from --data-dir):
  ose{YYYYMMDD}tp.csv          today's OSE theoretical-price file (IV + spot)
  {YYYYMMDD}open_interest.xlsx today's OI (sheet '別紙1')
  (optional) prior-day tp.csv + open_interest for the OI x IV sign (Convention B)

Output: greeks.json with, per expiry:
  - per_strike greeks (delta/gamma/vega/theta) at each strike's own IV
  - gex_A : GEX under the STANDARD dealer convention
            (dealers long calls / short puts)
  - gex_B : GEX under the OI x IV-INFERRED convention
            (per strike/side: OI up + IV up = customer BUY -> dealer SHORT;
             OI up + IV down/flat = customer SELL -> dealer LONG;
             ambiguous -> fall back to the standard sign, so B == A when
             no directional flow is detectable)
  - zero_gamma_A / zero_gamma_B : dealer gamma flip level, found by
            re-evaluating total dealer GEX across a spot grid
  - net_A / net_B : net dealer gamma / delta / vega / theta
  - top_gamma     : strikes with the largest |gamma x OI| (pin candidates)

Greeks use r = 0 (JPY approx), dividends folded into spot for short tenors.
Absolute yen GEX depends on the contract multiplier (default 1000 = large
N225 options); the PROFILE / sign / flip level are robust, exact yen is a
reference figure. Reuses parse_tp_csv()+build() from extract_iv.py.
"""
import argparse
import json
import math
import os
import re
from datetime import date

from extract_iv import parse_tp_csv, build

MULT = 1000          # N225 option multiplier (large). Mini = 100.
OI_EPS = 50          # min net OI change to call a directional flow
IV_EPS = 0.002       # min IV change (decimal, = 0.2 vol pts) for a flow read


# ----------------------------------------------------------------------------
# Black-Scholes greeks (r = 0)
# ----------------------------------------------------------------------------
def _np(x):
    return math.exp(-x * x / 2.0) / math.sqrt(2.0 * math.pi)


def _nc(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def greeks(S, K, T, sig, cp):
    """Return dict of delta/gamma/vega(per 1 vol pt)/theta(per day)."""
    if not (S > 0 and K > 0 and T > 0 and sig and sig > 0):
        return {'delta': 0.0, 'gamma': 0.0, 'vega': 0.0, 'theta': 0.0}
    srt = sig * math.sqrt(T)
    d1 = (math.log(S / K) + 0.5 * sig * sig * T) / srt
    delta = _nc(d1) if cp == 'C' else _nc(d1) - 1.0
    gamma = _np(d1) / (S * srt)
    vega = S * _np(d1) * math.sqrt(T) / 100.0
    theta = (-S * _np(d1) * sig / (2.0 * math.sqrt(T))) / 365.0
    return {'delta': delta, 'gamma': gamma, 'vega': vega, 'theta': theta}


def _gamma_at(S, K, T, sig):
    if not (S > 0 and K > 0 and T > 0 and sig and sig > 0):
        return 0.0
    srt = sig * math.sqrt(T)
    d1 = (math.log(S / K) + 0.5 * sig * sig * T) / srt
    return _np(d1) / (S * srt)


# ----------------------------------------------------------------------------
# Expiry / time helpers
# ----------------------------------------------------------------------------
def _sq_date(expiry_code):
    """expiry_code like '202607' -> 2nd-Friday SQ date."""
    y = int(expiry_code[:4]); m = int(expiry_code[4:6])
    d = date(y, m, 1)
    # weekday(): Mon=0..Sun=6 ; Friday=4
    first_fri = 1 + ((4 - d.weekday()) % 7)
    return date(y, m, first_fri + 7)


def _today_from_dir(data_dir):
    best = None
    for fn in os.listdir(data_dir):
        mo = re.search(r'(\d{8})', fn)
        if mo:
            d = mo.group(1)
            if best is None or d > best:
                best = d
    return best


# ----------------------------------------------------------------------------
# OI parsing (sheet 別紙1) -> {expiry: {strike: {'call_oi','put_oi'}}}
# ----------------------------------------------------------------------------
def parse_oi(path):
    from openpyxl import load_workbook
    out = {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb['別紙1']
    except Exception:
        return out
    pat_p = re.compile(r'NIKKEI\s*225\s*P\s*(\d{4})-(\d+)')
    pat_c = re.compile(r'NIKKEI\s*225\s*C\s*(\d{4})-(\d+)')

    def _num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0
    for row in ws.iter_rows(values_only=True):
        a = str(row[0]).strip() if row and row[0] else ''
        g = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        mp = pat_p.match(a)
        if mp and int(mp.group(2)) % 500 == 0:
            e = mp.group(1); k = int(mp.group(2))
            out.setdefault(e, {}).setdefault(k, {'call_oi': 0.0, 'put_oi': 0.0})
            out[e][k]['put_oi'] += _num(row[2] if len(row) > 2 else 0)
        mc = pat_c.match(g)
        if mc and int(mc.group(2)) % 500 == 0:
            e = mc.group(1); k = int(mc.group(2))
            out.setdefault(e, {}).setdefault(k, {'call_oi': 0.0, 'put_oi': 0.0})
            out[e][k]['call_oi'] += _num(row[8] if len(row) > 8 else 0)
    return out


def _find(data_dir, pattern, exclude=None):
    hits = []
    for fn in os.listdir(data_dir):
        if re.search(pattern, fn) and (not exclude or exclude not in fn):
            mo = re.search(r'(\d{8})', fn)
            if mo:
                hits.append((mo.group(1), os.path.join(data_dir, fn)))
    return sorted(hits)


# ----------------------------------------------------------------------------
# Convention B: dealer sign per strike/side from OI x IV flow
# ----------------------------------------------------------------------------
def _dealer_sign(oi_chg, iv_chg, side):
    """Return dealer position sign (+1 long / -1 short) for this option.
    Fallback (ambiguous) = standard convention: long call (+1), short put (-1).
    """
    fallback = +1 if side == 'C' else -1
    if oi_chg is None or iv_chg is None or oi_chg <= OI_EPS:
        return fallback, 'fallback'
    if iv_chg >= IV_EPS:          # customer BUYING -> dealer SHORT
        return -1, 'buy'
    if iv_chg <= -IV_EPS:         # customer SELLING -> dealer LONG
        return +1, 'sell'
    return fallback, 'fallback'   # OI up but IV flat -> ambiguous


# ----------------------------------------------------------------------------
# Snapshot history — lets B/B2 work even when the raw prior-day files are gone
# ----------------------------------------------------------------------------
def _load_history(path):
    try:
        with open(path, encoding='utf-8') as f:
            h = json.load(f)
        return h if isinstance(h, dict) else {}
    except Exception:
        return {}


def _save_history(path, history, keep=10):
    # keep only the most recent `keep` dates to bound file size
    dates = sorted(history.keys())
    for d in dates[:-keep]:
        history.pop(d, None)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, separators=(',', ':'))
    except Exception as e:
        print('[extract_greeks] WARNING: could not write history: %s' % e)


def _snapshot(oi_now, iv_expiries, spot=None, pct=0.18):
    """Compact per-day snapshot: OI (YYMM-keyed) + IV/ATM (YYYYMM-keyed).
    Kept small: only the analysed front expiries, strikes within ±pct of spot,
    integer OI and rounded IV. This keeps greeks_history.json lean.

    The window is deliberately wider than the one used for display: a strike
    must be present in YESTERDAY's snapshot for today's OI change to be
    computable, and on a 3% move a narrow window drops exactly the strikes
    that just became interesting."""
    def near(k):
        return spot is None or abs(int(k) - spot) <= spot * pct
    snap = {'oi': {}, 'iv': {}, 'atm_iv': {}}
    fronts = set()
    for e in iv_expiries:
        if len(e['expiry']) == 6:
            fronts.add(e['expiry'][2:])  # YYMM, to match OI keys
            snap['iv'][e['expiry']] = {str(int(p['strike'])): round(p['iv'], 4)
                                       for p in e['smile']
                                       if p.get('iv') and near(p['strike'])}
            if e.get('atm_iv') is not None:
                snap['atm_iv'][e['expiry']] = round(e['atm_iv'], 4)
    for ec4, strikes in (oi_now or {}).items():
        if fronts and ec4 not in fronts:
            continue  # keep only the front expiries we actually analyse
        snap['oi'][ec4] = {str(int(k)): {'call_oi': int((v or {}).get('call_oi', 0) or 0),
                                         'put_oi': int((v or {}).get('put_oi', 0) or 0)}
                           for k, v in strikes.items() if near(k)}
    return snap


def _regime_timeline(history, keep=10):
    """Recompute front-month net gamma (A/B/B2) per day from the snapshot
    history, so the card can show how the stability regime evolved over time."""
    dates = sorted(history.keys())
    out = []
    for i, d in enumerate(dates):
        snap = history[d] or {}
        iv = snap.get('iv') or {}
        oi = snap.get('oi') or {}
        atm = snap.get('atm_iv') or {}
        spot = snap.get('spot')
        if not iv or not oi:
            continue
        ecode = min(iv.keys())            # front expiry (YYYYMM)
        ec4 = ecode[2:]
        smile = iv.get(ecode) or {}
        oi_e = oi.get(ec4) or {}
        if not smile or not oi_e:
            continue
        if not spot:
            # estimate from the smile (min-IV strike sits near ATM/spot)
            cand = [(v, int(k)) for k, v in smile.items() if v]
            if not cand:
                continue
            spot = min(cand)[1]
        try:
            dd = date(int(d[:4]), int(d[4:6]), int(d[6:8]))
            T = max((_sq_date(ecode) - dd).days, 1) / 365.0
        except Exception:
            continue
        prior = history[dates[i - 1]] if i > 0 else None
        piv = ((prior or {}).get('iv', {}) or {}).get(ecode, {})
        poi_map = ((prior or {}).get('oi', {}) or {}).get(ec4, {})
        patm = ((prior or {}).get('atm_iv', {}) or {}).get(ecode)
        tatm = atm.get(ecode)
        atm_chg = (tatm - patm) if (tatm is not None and patm is not None) else None
        nA = nB = nB2 = 0.0
        for ks, sig in smile.items():
            K = int(ks)
            if K % 500 != 0 or abs(K - spot) > spot * 0.15:
                continue
            rec = oi_e.get(ks) or {}
            coi = rec.get('call_oi', 0) or 0
            poi_v = rec.get('put_oi', 0) or 0
            g = _gamma_at(spot, K, T, sig)
            nA += (coi - poi_v) * g
            prec = poi_map.get(ks) or {}
            coi_chg = (coi - (prec.get('call_oi', 0) or 0)) if prior else None
            poi_chg = (poi_v - (prec.get('put_oi', 0) or 0)) if prior else None
            iv_chg = (sig - piv[ks]) if (ks in piv) else None
            iv_rel = (iv_chg - atm_chg) if (iv_chg is not None and atm_chg is not None) else None
            scB, _ = _dealer_sign(coi_chg, iv_chg, 'C')
            spB, _ = _dealer_sign(poi_chg, iv_chg, 'P')
            nB += (scB * coi + spB * poi_v) * g
            scB2, _ = _dealer_sign(coi_chg, iv_rel, 'C')
            spB2, _ = _dealer_sign(poi_chg, iv_rel, 'P')
            nB2 += (scB2 * coi + spB2 * poi_v) * g
        out.append({'date': d, 'spot': spot,
                    'A': round(nA, 2), 'B': round(nB, 2), 'B2': round(nB2, 2)})
    return out[-keep:]


# ----------------------------------------------------------------------------
# Core build
# ----------------------------------------------------------------------------
def _mny_rel_chg(K, iv_now, spot_now, spot_was, atm_now, atm_was, iv_was_e):
    """Change in a strike's IV premium over ATM, measured at CONSTANT moneyness.

    Fixed-strike relative IV is contaminated when spot moves: on a down day a
    downside put's strike becomes relatively closer to the money, so its skew
    premium shrinks even with zero trading. Here we instead ask: at today's
    moneyness (K/spot_now), what premium did that same moneyness carry
    yesterday? The difference is the part attributable to actual flow.
    Returns None when any input is missing or the prior smile is too sparse.
    """
    if None in (iv_now, spot_now, spot_was, atm_now, atm_was) or not iv_was_e:
        return None
    if spot_now <= 0 or spot_was <= 0:
        return None
    k_equiv = (K / spot_now) * spot_was          # same moneyness, yesterday
    ks = sorted(iv_was_e)
    if not ks or k_equiv < ks[0] or k_equiv > ks[-1]:
        return None                               # outside prior smile: no basis
    # linear interpolation of yesterday's smile at the equivalent strike
    lo = max([x for x in ks if x <= k_equiv])
    hi = min([x for x in ks if x >= k_equiv])
    if hi == lo:
        iv_prev_equiv = iv_was_e[lo]
    else:
        w = (k_equiv - lo) / (hi - lo)
        iv_prev_equiv = iv_was_e[lo] * (1 - w) + iv_was_e[hi] * w
    return (iv_now - atm_now) - (iv_prev_equiv - atm_was)


def build_greeks(data_dir, max_expiries=3):
    today = _today_from_dir(data_dir)
    tps = _find(data_dir, r'ose\d{8}tp.*\.csv')
    ois = _find(data_dir, r'\d{8}open_interest.*\.xlsx')
    if not tps or not ois:
        return {'error': 'need today ose tp.csv and open_interest.xlsx',
                'have_tp': [d for d, _ in tps], 'have_oi': [d for d, _ in ois]}

    tp_today = tps[-1][1]
    oi_today = ois[-1][1]
    tp_prior = tps[-2][1] if len(tps) >= 2 else None
    oi_prior = ois[-2][1] if len(ois) >= 2 else None

    iv_expiries = build(parse_tp_csv(tp_today))
    spot = None
    for e in iv_expiries:
        if e.get('underlying'):
            spot = e['underlying']; break
    oi_now = parse_oi(oi_today)
    oi_was = parse_oi(oi_prior) if oi_prior else {}

    iv_prior_map = {}
    iv_prior_atm = {}
    spot_was = None
    if tp_prior:
        for e in build(parse_tp_csv(tp_prior)):
            iv_prior_map[e['expiry']] = {int(p['strike']): p['iv']
                                         for p in e['smile'] if p.get('iv')}
            iv_prior_atm[e['expiry']] = e.get('atm_iv')
            if spot_was is None and e.get('underlying'):
                spot_was = e['underlying']

    # --- Robust prior: fall back to the accumulated snapshot history when the
    # raw previous-day files are not in data/, so B/B2 keep working even if the
    # prior day's open_interest.xlsx / ose tp.csv were cleaned up. ---
    hist_path = os.path.join(data_dir, 'greeks_history.json')
    history = _load_history(hist_path)
    prior_src = 'raw' if (oi_prior and tp_prior) else None
    if not (oi_was and iv_prior_map):
        pdays = sorted(d for d in history if d < today)
        if pdays:
            ph = history[pdays[-1]]
            if not oi_was and ph.get('oi'):
                oi_was = {ec: {int(k): v for k, v in s.items()}
                          for ec, s in ph['oi'].items()}
            if not iv_prior_map and ph.get('iv'):
                iv_prior_map = {ec: {int(k): v for k, v in s.items()}
                                for ec, s in ph['iv'].items()}
                iv_prior_atm = dict(ph.get('atm_iv', {}))
            if spot_was is None:
                spot_was = ph.get('spot')
            prior_src = 'history:' + pdays[-1]

    prior_ok = bool(oi_was) and bool(iv_prior_map)
    out = {'as_of': today, 'spot': spot, 'r': 0.0, 'mult': MULT,
           'prior_used': prior_ok,
           'prior_src': prior_src,
           'prior_oi': (ois[-2][0] if len(ois) >= 2 else None),
           'prior_tp': (tps[-2][0] if len(tps) >= 2 else None),
           'conventions': {
               'A': 'standard: dealer long calls / short puts',
               'B': 'OI x IV inferred sign (fallback to A when ambiguous)',
               'B2': 'OI x RELATIVE-IV (strike IV chg minus ATM IV chg) — '
                     'isolates strike-specific demand from market-wide vol moves'},
           'expiries': []}

    liquid = [e for e in iv_expiries if len(e['expiry']) == 6][:max_expiries]
    for e in liquid:
        ecode = e['expiry']
        sq = _sq_date(ecode)
        T_days = max((sq - date(int(today[:4]), int(today[4:6]), int(today[6:8]))).days, 0)
        T = max(T_days, 0.5) / 365.0
        smile = {int(p['strike']): p['iv'] for p in e['smile'] if p.get('iv')}
        ec4 = ecode[2:]  # OI sheet uses YYMM (e.g. '2607'); iv uses YYYYMM
        oi_e = oi_now.get(ec4, {})
        oi_e_was = oi_was.get(ec4, {})
        iv_was_e = iv_prior_map.get(ecode, {})
        # market-wide ATM IV change for this expiry (Convention B2 baseline)
        atm_now = e.get('atm_iv')
        atm_was = iv_prior_atm.get(ecode)
        atm_iv_chg = (atm_now - atm_was) if (atm_now is not None and atm_was is not None) else None

        per = []
        net_A = {'gamma': 0.0, 'delta': 0.0, 'vega': 0.0, 'theta': 0.0}
        net_B = {'gamma': 0.0, 'delta': 0.0, 'vega': 0.0, 'theta': 0.0}
        net_B2 = {'gamma': 0.0, 'delta': 0.0, 'vega': 0.0, 'theta': 0.0}
        gex_A = []; gex_B = []; gex_B2 = []
        sign_meta = []
        # union of strikes that have either IV or OI, restricted to round
        # (500-point) strikes within +/-15% of spot. OI exists only at 500-pt
        # strikes; far strikes carry ~0 gamma, so this keeps the card lean
        # without affecting the relevant walls. (zero_gamma uses the full smile.)
        strikes = sorted(k for k in (set(smile) | set(oi_e))
                         if k % 500 == 0 and (not spot or abs(k - spot) <= spot * 0.15))
        for K in strikes:
            sig = smile.get(K) or (smile[min(smile, key=lambda x: abs(x - K))]
                                   if smile else None)
            if not sig:
                continue
            coi = (oi_e.get(K, {}) or {}).get('call_oi', 0.0)
            poi = (oi_e.get(K, {}) or {}).get('put_oi', 0.0)
            coi_w = (oi_e_was.get(K, {}) or {}).get('call_oi', 0.0)
            poi_w = (oi_e_was.get(K, {}) or {}).get('put_oi', 0.0)
            # A strike that was outside yesterday's window (spot moved a lot)
            # has no prior record. Subtracting a missing value from today's OI
            # would report the entire open interest as if it were built today —
            # badly overstating the flow on exactly the days that matter most.
            # Mark it unknown instead.
            had_prior_K = K in oi_e_was
            coi_chg = (coi - coi_w) if (oi_e_was and had_prior_K) else None
            poi_chg = (poi - poi_w) if (oi_e_was and had_prior_K) else None
            iv_chg = (sig - iv_was_e[K]) if (K in iv_was_e) else None
            # relative IV change = strike IV change minus market-wide ATM change
            iv_chg_rel = (iv_chg - atm_iv_chg) if (iv_chg is not None and atm_iv_chg is not None) else None
            # moneyness-adjusted version: compare the skew SPREAD at constant
            # moneyness instead of at a fixed strike. When spot moves a lot, a
            # fixed strike slides along the skew curve (a downside put becomes
            # relatively closer to ATM on a down day), which shrinks its IV
            # premium mechanically and would otherwise read as "sold".
            iv_chg_rel_mny = _mny_rel_chg(K, sig, spot, spot_was, atm_now,
                                          atm_was, iv_was_e)

            gc = greeks(spot, K, T, sig, 'C')
            gp = greeks(spot, K, T, sig, 'P')
            g = gc['gamma']  # same magnitude for C/P at K

            # dollar-gamma per 1% move
            unit = g * MULT * spot * spot * 0.01
            # Convention A
            a = (coi - poi) * unit
            # Convention B (absolute IV change)
            sc, rc = _dealer_sign(coi_chg, iv_chg, 'C')
            sp_, rp = _dealer_sign(poi_chg, iv_chg, 'P')
            b = (sc * coi + sp_ * poi) * unit
            # Convention B2 (relative IV change vs ATM)
            sc2, rc2 = _dealer_sign(coi_chg, iv_chg_rel, 'C')
            sp2, rp2 = _dealer_sign(poi_chg, iv_chg_rel, 'P')
            b2 = (sc2 * coi + sp2 * poi) * unit

            gex_A.append({'strike': K, 'gex': a / 1e8})   # 億円/1%
            gex_B.append({'strike': K, 'gex': b / 1e8})
            gex_B2.append({'strike': K, 'gex': b2 / 1e8})
            sign_meta.append({'strike': K, 'call': rc, 'put': rp,
                              'call_b2': rc2, 'put_b2': rp2})

            net_A['gamma'] += (coi - poi) * g
            net_B['gamma'] += (sc * coi + sp_ * poi) * g
            net_B2['gamma'] += (sc2 * coi + sp2 * poi) * g
            net_A['delta'] += coi * gc['delta'] - poi * gp['delta']
            net_B['delta'] += (sc * coi) * gc['delta'] + (sp_ * poi) * gp['delta']
            net_B2['delta'] += (sc2 * coi) * gc['delta'] + (sp2 * poi) * gp['delta']
            for kk in ('vega', 'theta'):
                net_A[kk] += (coi + poi) * gc[kk]
                net_B[kk] += (coi + poi) * gc[kk]
                net_B2[kk] += (coi + poi) * gc[kk]

            per.append({
                'strike': K, 'iv': round(sig, 5),
                'call_oi': int(coi), 'put_oi': int(poi),
                'call_oi_chg': (int(coi_chg) if coi_chg is not None else None),
                'put_oi_chg': (int(poi_chg) if poi_chg is not None else None),
                'iv_chg': (round(iv_chg, 5) if iv_chg is not None else None),
                'iv_chg_rel': (round(iv_chg_rel, 5) if iv_chg_rel is not None else None),
                'iv_chg_rel_mny': (round(iv_chg_rel_mny, 5) if iv_chg_rel_mny is not None else None),
                'call': {k: round(v, 6) for k, v in gc.items()},
                'put': {k: round(v, 6) for k, v in gp.items()},
            })

        zga = _zero_gamma(spot, T, smile, oi_e, oi_e_was, iv_was_e, 'A')
        zgb = _zero_gamma(spot, T, smile, oi_e, oi_e_was, iv_was_e, 'B')
        zgb2 = _zero_gamma(spot, T, smile, oi_e, oi_e_was, iv_was_e, 'B2',
                           atm_iv_chg=atm_iv_chg)
        top_gamma = sorted(
            [{'strike': p['strike'],
              'gamma_oi': round(p['call']['gamma'] * p['call_oi']
                                + p['put']['gamma'] * p['put_oi'], 4),
              'call_oi': p['call_oi'], 'put_oi': p['put_oi']} for p in per],
            key=lambda r: -r['gamma_oi'])[:6]

        out['expiries'].append({
            'expiry': ecode, 'label': '%d月限' % int(ecode[4:6]),
            'T_days': T_days, 'T': round(T, 5), 'spot': spot,
            'per_strike': per,
            'gex_A': gex_A, 'gex_B': gex_B, 'gex_B2': gex_B2,
            'zero_gamma_A': zga, 'zero_gamma_B': zgb, 'zero_gamma_B2': zgb2,
            'net_A': {k: round(v, 4) for k, v in net_A.items()},
            'net_B': {k: round(v, 4) for k, v in net_B.items()},
            'net_B2': {k: round(v, 4) for k, v in net_B2.items()},
            'atm_iv_chg': (round(atm_iv_chg, 5) if atm_iv_chg is not None else None),
            'top_gamma': top_gamma,
            'sign_meta': sign_meta,
            'has_oi_change': bool(oi_e_was),
            'has_iv_change': bool(iv_was_e),
        })

    # --- persist snapshots so future runs can recover the prior day without
    # the raw files. Save today, and (bootstrap) the raw prior if we had it. ---
    history[today] = _snapshot(oi_now, iv_expiries, spot)
    history[today]['spot'] = spot
    if oi_prior and tp_prior:
        prior_date = ois[-2][0]
        if prior_date not in history and oi_was and iv_prior_map:
            ps = {'oi': {}, 'iv': {}, 'atm_iv': {}}
            for ec4, strikes in oi_was.items():
                ps['oi'][ec4] = {str(int(k)): {'call_oi': (v or {}).get('call_oi', 0),
                                               'put_oi': (v or {}).get('put_oi', 0)}
                                 for k, v in strikes.items()}
            for ec, sm in iv_prior_map.items():
                ps['iv'][ec] = {str(int(k)): v for k, v in sm.items()}
            ps['atm_iv'] = {ec: a for ec, a in iv_prior_atm.items() if a is not None}
            history[prior_date] = ps
    _save_history(hist_path, history)
    out['regime_history'] = _regime_timeline(history)
    return out


def _zero_gamma(spot, T, smile, oi_e, oi_e_was, iv_was_e, conv, atm_iv_chg=None):
    """Find dealer gamma flip level by re-evaluating total GEX on a spot grid."""
    if not smile or not oi_e:
        return None
    lo, hi = spot * 0.90, spot * 1.10
    n = 81
    grid = [lo + (hi - lo) * i / (n - 1) for i in range(n)]

    def total_gex(Sx):
        tot = 0.0
        for K, sig in smile.items():
            coi = (oi_e.get(K, {}) or {}).get('call_oi', 0.0)
            poi = (oi_e.get(K, {}) or {}).get('put_oi', 0.0)
            g = _gamma_at(Sx, K, T, sig)
            if conv == 'A':
                tot += (coi - poi) * g
            else:
                coi_w = (oi_e_was.get(K, {}) or {}).get('call_oi', 0.0)
                poi_w = (oi_e_was.get(K, {}) or {}).get('put_oi', 0.0)
                ivc = (sig - iv_was_e[K]) if (K in iv_was_e) else None
                if conv == 'B2' and ivc is not None and atm_iv_chg is not None:
                    ivc = ivc - atm_iv_chg
                sc, _ = _dealer_sign(coi - coi_w if oi_e_was else None, ivc, 'C')
                sp_, _ = _dealer_sign(poi - poi_w if oi_e_was else None, ivc, 'P')
                tot += (sc * coi + sp_ * poi) * g
        return tot

    prev_S, prev_v = grid[0], total_gex(grid[0])
    crossings = []
    for Sx in grid[1:]:
        v = total_gex(Sx)
        if prev_v == 0 or (prev_v < 0) != (v < 0):
            if v != prev_v:
                cross = prev_S + (Sx - prev_S) * (0 - prev_v) / (v - prev_v)
            else:
                cross = Sx
            crossings.append(round(cross))
        prev_S, prev_v = Sx, v
    return {'flip': crossings[0] if crossings else None,
            'all_flips': crossings,
            'sign_at_spot': ('positive' if total_gex(spot) >= 0 else 'negative')}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--data-dir', default='data')
    ap.add_argument('--out', default='data/greeks.json')
    ap.add_argument('--max-expiries', type=int, default=3)
    a = ap.parse_args()
    res = build_greeks(a.data_dir, a.max_expiries)
    with open(a.out, 'w') as f:
        json.dump(res, f, ensure_ascii=False, separators=(',', ':'))
    if res.get('error'):
        print('[extract_greeks] ERROR:', res['error'], res)
    else:
        print('[extract_greeks] wrote %s | spot=%.0f | %d expiries'
              % (a.out, res['spot'] or 0, len(res['expiries'])))
        if not res.get('prior_used'):
            print('[extract_greeks] WARNING: no prior day available (raw files '
                  'or history) -> B and B2 fall back to A. This self-heals once '
                  'greeks_history.json has accumulated a prior day.')
        else:
            print('[extract_greeks] prior source: %s' % res.get('prior_src'))
        for e in res['expiries']:
            za = (e['zero_gamma_A'] or {}).get('flip')
            zb = (e['zero_gamma_B'] or {}).get('flip')
            zb2 = (e['zero_gamma_B2'] or {}).get('flip')
            print('  %s T=%dd  zeroGamma A=%s B=%s B2=%s  netGamma A=%.1f B=%.1f B2=%.1f'
                  % (e['label'], e['T_days'], za, zb, zb2,
                     e['net_A']['gamma'], e['net_B']['gamma'], e['net_B2']['gamma']))


if __name__ == '__main__':
    main()
