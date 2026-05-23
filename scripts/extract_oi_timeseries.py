#!/usr/bin/env python3
"""Extract multi-day OI timeseries from accumulated `*open_interest.xlsx` files.

Reads all `*open_interest.xlsx` files in `data/` (sorted by YYYYMMDD prefix),
extracts per-day:
  - Futures: total OI per market (大/mini/TOPIX) per limgetsu
  - Options aggregate: total Put/Call OI per limgetsu
  - Options per strike: per-strike OI for key high-OI strikes

Writes `data/oi_timeseries.json`. Standalone module — does not modify
extract.py state. Safe to run repeatedly.
"""
import os
import re
import json
import glob
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print('[oi_timeseries] ERROR: openpyxl not installed', file=sys.stderr)
    sys.exit(1)


def safe_num(v, default=0):
    if v is None or v == '':
        return default
    try:
        return float(str(v).replace(',', ''))
    except (ValueError, TypeError):
        return default


def round500(x):
    return round(x / 500) * 500


# --- Futures section (mirrors extract.py extract_s02 logic) -----------------

def extract_futures_oi(wb_oi):
    """Return {market: {total_oi, by_limgetsu: {limgetsu_label: oi}}}."""
    if 'デリバティブ建玉残高状況' not in wb_oi.sheetnames:
        return {}
    ws = wb_oi['デリバティブ建玉残高状況']
    result = {'nk225_large': {'by_limgetsu': {}}, 'nk225_mini': {'by_limgetsu': {}}, 'topix': {'by_limgetsu': {}}}

    # Find 指数先物 section boundaries
    fut_start, fut_end = None, None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        if '指数先物取引' in a_val:
            fut_start = row[0].row
        elif fut_start and ('商品先物' in a_val or '国債先物オプション' in a_val or '指数オプション' in a_val):
            fut_end = row[0].row
            break
    if not fut_start:
        return result
    if not fut_end:
        fut_end = min(fut_start + 80, ws.max_row)

    # LEFT side (A-F) — 日経225 large, then TOPIX
    left_section = None
    for row in ws.iter_rows(min_row=fut_start, max_row=fut_end, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        b_val = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''

        if '日経225' in a_val and 'mini' not in a_val.lower() and 'ミニ' not in a_val and 'マイクロ' not in a_val and 'オプション' not in a_val:
            left_section = 'nk225_large'
        elif 'TOPIX' in a_val and 'ミニ' not in a_val:
            left_section = 'topix'
        elif a_val and left_section and '日経' not in a_val and 'TOPIX' not in a_val:
            if left_section == 'topix':
                left_section = None
        if not left_section:
            continue

        if '合計' in b_val:
            result[left_section]['total_oi'] = safe_num(row[3].value if len(row) > 3 else None)
            if left_section == 'nk225_large':
                left_section = None
        elif b_val and ('月限' in b_val or '年' in b_val):
            oi = safe_num(row[3].value if len(row) > 3 else None)
            result[left_section]['by_limgetsu'][b_val] = oi

    # RIGHT side (H-M) — 日経225mini
    right_section = None
    for row in ws.iter_rows(min_row=fut_start, max_row=fut_end, values_only=False):
        h_val = str(row[7].value).strip() if len(row) > 7 and row[7].value else ''
        i_val = str(row[8].value).strip() if len(row) > 8 and row[8].value else ''

        if '日経225mini' in h_val or '日経225ミニ' in h_val:
            right_section = 'nk225_mini'
        elif '日経225マイクロ' in h_val or 'ミニTOPIX' in h_val:
            right_section = None
        if right_section != 'nk225_mini':
            continue

        if '合計' in i_val:
            result['nk225_mini']['total_oi'] = safe_num(row[10].value if len(row) > 10 else None)
            right_section = None
        elif i_val and ('月限' in i_val or '年' in i_val):
            oi = safe_num(row[10].value if len(row) > 10 else None)
            result['nk225_mini']['by_limgetsu'][i_val] = oi

    return result


# --- Options section (mirrors extract.py extract_s06 logic, no range limit) -

PAT_PUT  = re.compile(r'NIKKEI\s*225\s*P\s*(\d{4})-(\d+)')
PAT_CALL = re.compile(r'NIKKEI\s*225\s*C\s*(\d{4})-(\d+)')


def extract_options_oi(wb_oi):
    """Return per-day options OI structure:
      {
        'aggregate': {expiry_yymm: {'put_total': X, 'call_total': Y}},
        'per_strike': {expiry_yymm: {strike: {'put_oi': X, 'call_oi': Y}}}
      }
    """
    if '別紙1' not in wb_oi.sheetnames:
        return {'aggregate': {}, 'per_strike': {}}
    ws = wb_oi['別紙1']

    aggregate = defaultdict(lambda: {'put_total': 0.0, 'call_total': 0.0})
    per_strike = defaultdict(lambda: defaultdict(lambda: {'put_oi': 0.0, 'call_oi': 0.0}))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''

        m = PAT_PUT.match(a_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            oi = safe_num(row[2].value if len(row) > 2 else None)
            if strike % 500 == 0:  # restrict to 500-yen grid for cleaner timeseries
                aggregate[expiry]['put_total'] += oi
                per_strike[expiry][strike]['put_oi'] += oi

        m = PAT_CALL.match(g_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            oi = safe_num(row[8].value if len(row) > 8 else None)
            if strike % 500 == 0:
                aggregate[expiry]['call_total'] += oi
                per_strike[expiry][strike]['call_oi'] += oi

    # Convert defaultdicts → plain dicts for JSON.
    # IMPORTANT: strike keys are stringified so they survive JSON round-trips
    # (JSON object keys are always strings; this keeps lookups consistent
    # whether the snapshot was freshly extracted or loaded from disk).
    return {
        'aggregate': {k: dict(v) for k, v in aggregate.items()},
        'per_strike': {ek: {str(sk): dict(sv) for sk, sv in inner.items()}
                       for ek, inner in per_strike.items()},
    }


# --- Main aggregation ------------------------------------------------------

def extract_oi_timeseries(data_dir, max_days=20, top_strikes=8, nearest_expiry_only=True,
                          out_path=None):
    """Build the multi-day OI timeseries dict — INCREMENTAL/ACCUMULATING.

    The history is persisted inside the output JSON itself (under '_snapshots'),
    so it grows by one data point per pipeline run and survives even when old
    `*open_interest.xlsx` files are deleted/overwritten in data/.

    Args:
      data_dir: directory containing `*open_interest.xlsx` daily files.
      max_days: keep at most this many most-recent business days.
      top_strikes: number of top-OI strikes to track individually (PER expiry).
      nearest_expiry_only: if True, limit top_puts/top_calls to the nearest
        major expiry only (e.g. 06月限 only, no 07月限 mixing). Default True.
      out_path: path to the existing oi_timeseries.json (to load prior history).

    Returns dict ready to be serialized as oi_timeseries.json.
    """
    # --- 1. Load prior accumulated snapshots from the existing JSON ----------
    #   day_data[date_str] = {'futures': {...}, 'options': {...}}
    day_data = {}
    if out_path and os.path.exists(out_path):
        try:
            with open(out_path, 'r', encoding='utf-8') as f:
                prev = json.load(f)
            prev_snaps = prev.get('_snapshots', {})
            if isinstance(prev_snaps, dict):
                day_data.update(prev_snaps)
                print('[oi_timeseries] loaded %d prior snapshots from %s' % (len(prev_snaps), out_path))
        except Exception as e:
            print('[oi_timeseries] WARN: could not load prior history: %s' % e)

    # --- 2. Extract snapshots from any Excel files present, merge in ---------
    pattern = os.path.join(data_dir, '*open_interest.xlsx')
    file_candidates = []
    for fp in glob.glob(pattern):
        m = re.search(r'(\d{8})', os.path.basename(fp))
        if m:
            file_candidates.append((m.group(1), fp))
    file_candidates.sort(key=lambda x: x[0])

    new_dates = []
    for date_str, fp in file_candidates:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            fut = extract_futures_oi(wb)
            opt = extract_options_oi(wb)
            day_data[date_str] = {'futures': fut, 'options': opt}  # overwrite/add
            new_dates.append(date_str)
        except Exception as e:
            print('[oi_timeseries] WARN: parse failed on %s: %s' % (fp, e))
            continue

    if not day_data:
        return {'error': 'No snapshots: neither prior history nor *open_interest.xlsx in %s' % data_dir}

    if new_dates:
        print('[oi_timeseries] merged %d Excel file(s): %s' % (len(new_dates), new_dates))

    # --- 3. Trim to the most-recent max_days dates ---------------------------
    all_dates_sorted = sorted(day_data.keys())
    if len(all_dates_sorted) > max_days:
        keep = set(all_dates_sorted[-max_days:])
        day_data = {d: v for d, v in day_data.items() if d in keep}
        all_dates_sorted = sorted(day_data.keys())
    dates = all_dates_sorted
    print('[oi_timeseries] total accumulated days: %d (%s 〜 %s)' % (
        len(dates), dates[0] if dates else '-', dates[-1] if dates else '-'))

    # --- Build futures timeseries ------------------------------------------
    futures_out = {}
    for market in ('nk225_large', 'nk225_mini', 'topix'):
        # Collect ALL limgetsu labels seen across days for this market
        all_limgetsu = set()
        for d in dates:
            dd = day_data.get(d, {}).get('futures', {}).get(market, {})
            all_limgetsu.update(dd.get('by_limgetsu', {}).keys())

        total_series = []
        by_limgetsu = {lim: [] for lim in all_limgetsu}
        for d in dates:
            dd = day_data.get(d, {}).get('futures', {}).get(market, {})
            total_series.append(dd.get('total_oi', 0))
            for lim in all_limgetsu:
                by_limgetsu[lim].append(dd.get('by_limgetsu', {}).get(lim, 0))

        # Sort limgetsu by label (nearest first)
        sorted_lims = sorted(all_limgetsu)
        futures_out[market] = {
            'total': total_series,
            'by_limgetsu': {lim: by_limgetsu[lim] for lim in sorted_lims},
            'limgetsu_order': sorted_lims,
        }

    # --- Build options aggregate timeseries --------------------------------
    all_expiries = set()
    for d in dates:
        all_expiries.update(day_data.get(d, {}).get('options', {}).get('aggregate', {}).keys())

    options_aggregate = {}
    for expiry in sorted(all_expiries):
        put_series, call_series = [], []
        for d in dates:
            agg = day_data.get(d, {}).get('options', {}).get('aggregate', {}).get(expiry, {})
            put_series.append(agg.get('put_total', 0))
            call_series.append(agg.get('call_total', 0))
        # Skip expiries with negligible OI (e.g. far months, expired weeklies)
        if max(put_series + call_series) < 100:
            continue
        # Convert '2606' -> '2026年06月限'
        label = '20%s年%s月限' % (expiry[:2], expiry[2:]) if len(expiry) == 4 else expiry
        options_aggregate[expiry] = {
            'label': label,
            'put_total': put_series,
            'call_total': call_series,
        }

    # --- Pick top strikes by latest OI, SEPARATED into Puts and Calls ------
    # Puts and Calls have very different OI distributions: deep OTM puts get
    # huge hedge flows while calls cluster near ATM. Mixing them in one
    # ranking causes calls to be crowded out. Track them separately.
    #
    # If nearest_expiry_only=True (default), only consider the nearest major
    # expiry (smallest YYMM value seen) to avoid mixing 06月限 with 07月限.
    top_puts = []
    top_calls = []
    latest_date = dates[-1] if dates else None
    if latest_date:
        latest_opt = day_data.get(latest_date, {}).get('options', {}).get('per_strike', {})

        # Determine which expiries to consider
        if nearest_expiry_only and latest_opt:
            # Pick the expiry with the largest total OI (= the "near major" expiry).
            # Using total OI is more robust than just smallest YYMM because the
            # absolutely-nearest may be a fading weekly with little OI.
            expiry_scores = []
            for ek, strikes in latest_opt.items():
                total = 0
                for s, v in strikes.items():
                    total += v.get('put_oi', 0) + v.get('call_oi', 0)
                expiry_scores.append((total, ek))
            expiry_scores.sort(reverse=True)
            target_expiries = [expiry_scores[0][1]] if expiry_scores else []
        else:
            target_expiries = list(latest_opt.keys())

        # Score all P / C strikes across target_expiries
        put_scored, call_scored = [], []
        for expiry in target_expiries:
            for strike, vals in latest_opt.get(expiry, {}).items():
                p, c = vals.get('put_oi', 0), vals.get('call_oi', 0)
                if p >= 500:
                    put_scored.append((p, expiry, strike))
                if c >= 500:
                    call_scored.append((c, expiry, strike))

        put_scored.sort(reverse=True)
        call_scored.sort(reverse=True)

        def build_strike_series(scored, typ, target_list):
            for score, expiry, strike in scored[:top_strikes]:
                history = []
                for d in dates:
                    dd = day_data.get(d, {}).get('options', {}).get('per_strike', {}).get(expiry, {})
                    s_data = dd.get(strike, {}) if isinstance(dd, dict) else {}
                    key = 'put_oi' if typ == 'P' else 'call_oi'
                    history.append(s_data.get(key, 0))
                strike_int = int(strike)
                label_exp = '20%s年%s月限' % (expiry[:2], expiry[2:]) if len(expiry) == 4 else expiry
                target_list.append({
                    'label':       '%s %s%d' % (label_exp, typ, strike_int),
                    'short_label': '%s%d' % (typ, strike_int),  # no limgetsu suffix when all-same-expiry
                    'expiry':      expiry,
                    'strike':      strike_int,
                    'type':        typ,
                    'current_oi':  history[-1] if history else 0,
                    'oi_history':  history,
                })

        build_strike_series(put_scored,  'P', top_puts)
        build_strike_series(call_scored, 'C', top_calls)

    output = {
        'dates': dates,
        'n_dates': len(dates),
        'futures': futures_out,
        'options': {
            'aggregate':  options_aggregate,
            'top_puts':   top_puts,
            'top_calls':  top_calls,
            'top_expiry': target_expiries[0] if (latest_date and target_expiries) else None,
        },
        'generated_at': dates[-1] if dates else '',
        # Persist raw per-day snapshots so history accumulates across runs even
        # when old *open_interest.xlsx files are removed from data/.
        '_snapshots': day_data,
    }
    return output


# --- CLI entry point -------------------------------------------------------

def main():
    import argparse
    ap = argparse.ArgumentParser(description='Extract daily OI timeseries')
    ap.add_argument('--data-dir',   default='data', help='dir with *open_interest.xlsx')
    ap.add_argument('--out',        default='data/oi_timeseries.json', help='output JSON path')
    ap.add_argument('--max-days',   type=int, default=20)
    ap.add_argument('--top-strikes', type=int, default=8,
                    help='number of top strikes per expiry to track individually')
    ap.add_argument('--all-expiries', action='store_true',
                    help='include strikes from ALL expiries (default: only the nearest major expiry by OI)')
    args = ap.parse_args()

    result = extract_oi_timeseries(args.data_dir, max_days=args.max_days,
                                   top_strikes=args.top_strikes,
                                   nearest_expiry_only=not args.all_expiries,
                                   out_path=args.out)

    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    if 'error' in result:
        print('[oi_timeseries] %s' % result['error'])
        return 0

    n_puts  = len(result['options']['top_puts'])
    n_calls = len(result['options']['top_calls'])
    n_expiries = len(result['options']['aggregate'])
    top_exp = result['options'].get('top_expiry')
    scope = ('expiry=%s' % top_exp) if top_exp else 'all-expiries'
    print('[oi_timeseries] wrote %s | %d days | %d futures markets | %d expiries | %d top puts / %d top calls (%s)'
          % (args.out, result['n_dates'], 3, n_expiries, n_puts, n_calls, scope))
    return 0


if __name__ == '__main__':
    sys.exit(main())
