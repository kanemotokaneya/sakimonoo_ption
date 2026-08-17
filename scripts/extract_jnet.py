#!/usr/bin/env python3
"""
extract_jnet.py — per-broker J-NET (off-auction / cross) futures volume.

Parses `<date>_volume_by_participant_whole_day_J-NET.xlsx` and aggregates the
day-session J-NET volume by broker and product class (large / mini futures,
topix, options). Accumulates a daily history in `jnet_history.json` so broker
cross activity (e.g. UBS) can be tracked over time and eyeballed against price.

IMPORTANT: J-NET data has NO buy/sell split. This is VOLUME only, not
direction. Use it as a hypothesis ("did a broker's cross spike coincide with a
price stall?") to confirm against subsequent price action — never as a direct
buy/sell signal.
"""
import argparse
import glob
import json
import os
import re

import openpyxl

PRODUCTS = {'NK225F': 'large', 'NK225MF': 'mini',
            'TOPIXF': 'topix', 'NK225E': 'option'}

# Broker classification (mirrors extract.py PARTICIPANT_RULES)
_PART_RULES = [
    ('us', ['ゴールドマン', 'ＪＰモルガン', 'シティ', 'ビーオブエー', 'モルガン']),
    ('eu', ['ＵＢＳ', 'ソシエテ', 'バークレイズ', 'ＨＳＢＣ', 'ドイツ', 'ナティクシス']),
    ('hf', ['ＡＢＮ', 'ＢＮＰ', 'サスケハナ', 'インタラクティブ', 'フィリップ']),
    ('domestic', ['野村', '大和', 'みずほ', '三菱', 'ＳＭＢＣ', 'ＳＢＩ', '楽天',
                  '松井', '岩井', '豊証券', '立花', 'むさし', '日産', '岡三',
                  '安藤', '光世', '東海東京', '広田', '三田', 'マネックス']),
]
_CAT_LABEL = {'us': '米系', 'eu': '欧系', 'hf': 'HF代理', 'domestic': '国内', 'other': 'その他'}


def _classify(name):
    s = str(name or '')
    for cat, kws in _PART_RULES:
        for kw in kws:
            if kw in s:
                return cat
    return 'other'


def _parse_opt_name(prod):
    """'NIKKEI 225 OOP C2609-70500' -> ('C', '2609', 70500).

    The expiry code matters: the same strike in different expiries is a
    different contract, and merging them would invent crosses that never
    happened.
    """
    m = re.search(r'([CP])(\d{4})-(\d+)', str(prod or ''))
    if not m:
        return None, None, None
    return m.group(1), m.group(2), int(m.group(3))


def _broker_opt_pc(path):
    """Per-broker put/call option-cross lots today, from all NK225E rows."""
    import collections
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['手口上位一覧'] if '手口上位一覧' in wb.sheetnames else wb.worksheets[0]
    out = collections.defaultdict(lambda: {'p': 0, 'c': 0})
    for r in ws.iter_rows(values_only=True):
        if not r or str(r[0]).strip() != 'NK225E':
            continue
        side, _, _ = _parse_opt_name(r[2] if len(r) > 2 else None)
        broker = str(r[5]).strip() if len(r) > 5 and r[5] else None
        vol = r[7] if len(r) > 7 and isinstance(r[7], (int, float)) else None
        if side and broker and vol:
            out[broker]['p' if side == 'P' else 'c'] += int(vol)
    return dict(out)


def parse_option_crosses(path, min_vol=100, keep=8):
    """Detect notable J-NET option blocks, grouped by (expiry, side, strike).

    A "cross" shows up as two brokers with large, near-equal volume at the same
    contract (e.g. みずほ600 / ＡＢＮ600). We surface these dynamically — whoever
    the counterparties are that day — and flag the domestic<->overseas blocks.

    Note on `size`: J-NET volume is reported per participant, and a matched
    cross is the same contracts counted on both sides. We therefore report the
    larger leg as the block size (not the sum), which is the number of
    contracts that actually changed hands.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['手口上位一覧'] if '手口上位一覧' in wb.sheetnames else wb.worksheets[0]
    groups = {}   # (expiry, side, strike) -> list[(broker, vol)]
    for r in ws.iter_rows(values_only=True):
        if not r or str(r[0]).strip() != 'NK225E':
            continue
        side, expiry, strike = _parse_opt_name(r[2] if len(r) > 2 else None)
        if side is None:
            continue
        broker = str(r[5]).strip() if len(r) > 5 and r[5] else None
        vol = r[7] if len(r) > 7 and isinstance(r[7], (int, float)) else None
        if broker and vol:
            groups.setdefault((expiry, side, strike), []).append((broker, float(vol)))
    crosses = []
    for (expiry, side, strike), legs in groups.items():
        legs.sort(key=lambda x: -x[1])
        top = legs[0][1]
        if top < min_vol:
            continue
        # counterparty = second broker within 30% of the top volume
        matched = len(legs) > 1 and legs[1][1] >= top * 0.7
        named = [{'broker': b, 'vol': v, 'cat': _classify(b),
                  'cat_label': _CAT_LABEL[_classify(b)]}
                 for b, v in legs if v >= min_vol * 0.5][:4]
        cats = {l['cat'] for l in named}
        dom_vs_ovs = ('domestic' in cats) and bool(cats & {'us', 'eu', 'hf'})
        crosses.append({
            'side': side, 'strike': strike, 'expiry': expiry,
            'size': round(top), 'total': round(sum(v for _, v in legs)),
            'is_cross': matched, 'domestic_vs_overseas': dom_vs_ovs,
            'legs': named,
        })
    # rank: domestic<->overseas crosses first, then by size
    crosses.sort(key=lambda c: (not c['domestic_vs_overseas'], not c['is_cross'], -c['size']))
    return crosses[:keep]


def _digits(s):
    return re.sub(r'\D', '', str(s or ''))


def parse_jnet(path):
    """Return (date_str, {broker: {large,mini,topix,option}})."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['手口上位一覧'] if '手口上位一覧' in wb.sheetnames else wb.worksheets[0]
    date = None
    brokers = {}
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        # trading date line
        if len(r) > 2 and r[1] and '取引日' in str(r[1]) and r[2]:
            d = _digits(r[2])
            if len(d) >= 8:
                date = d[:8]
        pc = str(r[0]).strip() if r[0] else ''
        if pc in PRODUCTS:
            broker = str(r[5]).strip() if len(r) > 5 and r[5] else None
            vol = r[7] if len(r) > 7 and isinstance(r[7], (int, float)) else None
            if broker and vol is not None:
                d = brokers.setdefault(broker, {'large': 0.0, 'mini': 0.0,
                                                'topix': 0.0, 'option': 0.0})
                d[PRODUCTS[pc]] += float(vol)
    if not date:
        m = re.search(r'(\d{8})', os.path.basename(path))
        date = m.group(1) if m else None
    return date, brokers


def _load(path):
    try:
        with open(path, encoding='utf-8') as f:
            h = json.load(f)
        return h if isinstance(h, dict) else {}
    except Exception:
        return {}


def _save(path, history, keep=40):
    for d in sorted(history.keys())[:-keep]:
        history.pop(d, None)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, separators=(',', ':'))


def _spot_from_dir(data_dir):
    """Best-effort: read spot from greeks.json or data.json in the dir."""
    for name in ('greeks.json', 'data.json'):
        p = os.path.join(data_dir, name)
        try:
            d = json.load(open(p, encoding='utf-8'))
            if d.get('spot'):
                return d['spot']
            sp = (d.get('meta') or d.get('metadata') or {}).get('nikkei')
            if sp:
                return sp
        except Exception:
            pass
    return None


def build_jnet(data_dir, spot=None):
    files = sorted(glob.glob(os.path.join(
        data_dir, '*volume_by_participant_whole_day_J-NET*.xlsx')))
    if not files:
        return {'error': 'no J-NET file found'}
    date, brokers = parse_jnet(files[-1])
    if spot is None:
        spot = _spot_from_dir(data_dir)

    # per-broker put/call option-cross lots today (from NK225E rows)
    opt_pc = _broker_opt_pc(files[-1])

    rows = []
    for b, d in brokers.items():
        fut = d['large'] + d['mini']
        pc = opt_pc.get(b, {'p': 0, 'c': 0})
        rows.append({'broker': b, 'large': d['large'], 'mini': d['mini'],
                     'topix': d['topix'], 'option': d['option'], 'fut': fut,
                     'opt_p': pc['p'], 'opt_c': pc['c']})
    rows.sort(key=lambda x: -x['fut'])

    out = {'date': date, 'spot': spot, 'brokers': rows}

    # notable option cross blocks by strike (dynamic — whoever moved big today)
    try:
        out['option_crosses'] = parse_option_crosses(files[-1])
    except Exception:
        out['option_crosses'] = []

    # accumulate history (compact: broker -> fut/large/mini/option + put/call flow)
    hist_path = os.path.join(data_dir, 'jnet_history.json')
    history = _load(hist_path)
    history[date] = {
        'spot': spot,
        'brokers': {r['broker']: {'fut': r['fut'], 'large': r['large'],
                                  'mini': r['mini'], 'opt': r['option'],
                                  'opt_p': r['opt_p'], 'opt_c': r['opt_c']}
                    for r in rows},
        'opt_crosses': [{'side': c['side'], 'strike': c['strike'],
                         'size': c['size'],
                         'top': (c['legs'][0]['broker'] if c['legs'] else ''),
                         'dvo': c['domestic_vs_overseas']}
                        for c in out['option_crosses'][:5]],
    }
    _save(hist_path, history)
    out['history_dates'] = sorted(history.keys())
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--data-dir', default='data')
    ap.add_argument('--out', default='data/jnet.json')
    ap.add_argument('--spot', type=float, default=None)
    a = ap.parse_args()
    res = build_jnet(a.data_dir, spot=a.spot)
    with open(a.out, 'w', encoding='utf-8') as f:
        json.dump(res, f, ensure_ascii=False, indent=2)
    if res.get('error'):
        print('[extract_jnet] %s' % res['error'])
        return
    ubs = next((r for r in res['brokers']
                if 'ＵＢＳ' in r['broker'] or 'UBS' in r['broker']), None)
    print('[extract_jnet] wrote %s | date=%s | %d brokers | spot=%s'
          % (a.out, res['date'], len(res['brokers']), res['spot']))
    if ubs:
        print('[extract_jnet]   UBS futures J-NET: large=%.0f mini=%.0f total=%.0f'
              % (ubs['large'], ubs['mini'], ubs['fut']))
    print('[extract_jnet]   history dates: %s' % res['history_dates'])


if __name__ == '__main__':
    main()
